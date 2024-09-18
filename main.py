import pandas as pd
from datetime import datetime, timedelta
import requests
import json
import pytz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

load_dotenv()

def get_token(client_id, client_secret, tenant_id):
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    headers = {
        'content-Type': 'application/x-www-form-urlencoded',
        'Accept': '*/*'
    }
    data = {
        'client_id': client_id,
        'scope': 'https://graph.microsoft.com/.default',
        'client_secret': client_secret,
        'grant_type': 'client_credentials'
    }
    response = requests.post(url, headers=headers, data=data)
    if response.status_code != 200:
        print(f"Error getting token: {response.status_code}")
        print(response.text)
        return None
    return response.json().get('access_token')

def get_call_logs(token, from_date, to_date):
    url = f"https://graph.microsoft.com/v1.0/communications/callRecords/getPstnCalls(fromDateTime={from_date},toDateTime={to_date})"
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json'
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print(f"Error getting call logs: {response.status_code}")
        print(response.text)
        return None
    return response.json()

def send_email(subject, body, recipient, attachment_paths):
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))  # Default to port 587 for TLS
    smtp_username = os.getenv('SMTP_USERNAME')
    smtp_password = os.getenv('SMTP_PASSWORD')
    use_tls = os.getenv('USE_TLS', 'True').lower() == 'true'  # Default to using TLS
    msg = MIMEMultipart()
    msg['From'] = smtp_username
    msg['To'] = recipient
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    for attachment_path in attachment_paths:
        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
            msg.attach(part)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        if use_tls:
            server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)
        
def generate_excel_report(df, excel_filename):
    # Create a copy of the DataFrame to avoid modifying the original
    report_df = df.copy()

    # Convert UTC times to Europe/London time zone
    london_tz = pytz.timezone('Europe/London')

    # Select and rename columns (only if they exist)
    if 'startDateTime' in report_df.columns:
        report_df.loc[:, 'Start Time'] = pd.to_datetime(report_df['startDateTime']).dt.tz_convert(london_tz)
        report_df.loc[:, 'Call Date'] = report_df['Start Time'].dt.strftime('%d-%m-%y')
        report_df.loc[:, 'Call Time'] = report_df['Start Time'].dt.strftime('%H:%M')

    if 'endDateTime' in report_df.columns:
        report_df.loc[:, 'End Time'] = pd.to_datetime(report_df['endDateTime']).dt.tz_convert(london_tz)

    if 'userDisplayName' in report_df.columns:
        report_df.loc[:, 'User Display Name'] = report_df['userDisplayName']

    if 'userPrincipalName' in report_df.columns:
        report_df.loc[:, 'User Principal Name'] = report_df['userPrincipalName']

    if 'callerNumber' in report_df.columns:
        report_df.loc[:, 'Caller ID'] = report_df['callerNumber']

    if 'calleeNumber' in report_df.columns:
        report_df.loc[:, 'Destination'] = report_df['calleeNumber']

    if 'callType' in report_df.columns:
        report_df.loc[:, 'Call Type'] = report_df['callType'].str.capitalize()

    if 'duration' in report_df.columns:
        # Convert duration to minutes directly
        report_df.loc[:, 'Talking'] = report_df['duration'].astype(float) / 60
        report_df.loc[:, 'Totals'] = report_df['Talking']

    if 'charge' in report_df.columns:
        report_df.loc[:, 'Cost'] = report_df['charge']

    # Add 'Status' column if 'Talking' exists
    if 'Talking' in report_df.columns:
        report_df.loc[:, 'Status'] = report_df['Talking'].apply(lambda x: 'Answered' if x > 1 else 'Unanswered')

    # Round numeric columns
    numeric_columns = ['Talking', 'Totals', 'Cost']
    for col in numeric_columns:
        if col in report_df.columns:
            report_df.loc[:, col] = report_df[col].round(2)

    # Add a blank column
    report_df.loc[:, ''] = ''

    # Select and order final columns (only if they exist)
    final_columns = ['', '', 'User Display Name', 'User Principal Name', 'Call Date', 'Call Time', 
                     'Caller ID', 'Destination', 'Call Type', 'Status', 'Talking', 'Totals', 'Cost']
    report_df = report_df[[col for col in final_columns if col in report_df.columns]]

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Call Data'

    # Write the data to the sheet
    for r in dataframe_to_rows(report_df, index=False, header=True):
        ws.append(r)

    # Format the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Calculate and add total row
    total_calls = len(df)
    total_talking = df['Talking'].sum() if 'Talking' in df.columns else 0
    total_cost = df['Cost'].sum() if 'Cost' in df.columns else 0
    total_row = ['Total:', '', '', '', f'Calls - {total_calls}', '', '', '', '', '', f'{total_talking:.2f}', f'{total_talking:.2f}', f'{total_cost:.2f}']
    ws.append(total_row)

    # Format total row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Save the Excel file
    wb.save(excel_filename)
    print(f"Call data report generated: {excel_filename}")

# Set your Azure AD app details
client_id = os.getenv("CLIENT_ID")
client_secret = os.getenv("CLIENT_SECRET")
tenant_id = os.getenv("TENANT_ID")

# Get the list of userPrincipalNames from the .env file
user_principal_names = os.getenv("USER_PRINCIPAL_NAMES", "").split(",")
manager_email = os.getenv("MANAGER_EMAIL")

# Get date range (last 7 days)
london_tz = pytz.timezone('Europe/London')
end_date = datetime.now(london_tz)
start_date = end_date - timedelta(days=7)
from_date = start_date.astimezone(pytz.utc).isoformat()
to_date = end_date.astimezone(pytz.utc).isoformat()

print(f"Querying for calls from {from_date} to {to_date}")

# Get the token and call logs
token = get_token(client_id, client_secret, tenant_id)
if not token:
    print("Failed to get token. Exiting.")
    exit(1)

all_call_logs = get_call_logs(token, from_date, to_date)

if not all_call_logs:
    print("No call data found for the specified date range.")
    exit(0)

# Process the call logs
df = pd.DataFrame(all_call_logs['value'])
filtered_df = df[df['userPrincipalName'].isin(user_principal_names)]

# Format start and end dates as strings
start_date_str = start_date.strftime('%Y%m%d')
end_date_str = end_date.strftime('%Y%m%d')

# Generate the consolidated report for the manager
consolidated_excel_filename = f'consolidated_calls_report_{start_date_str}_{end_date_str}.xlsx'
generate_excel_report(filtered_df, consolidated_excel_filename)

# Initialize a list to store users without call data
users_without_data = []

# Generate and send individual reports to each user
for user_principal_name in user_principal_names:
    # Filter the call logs for the current user
    user_df = df[df['userPrincipalName'] == user_principal_name]
    
    if user_df.empty:
        print(f"No call data found for user: {user_principal_name}")
        users_without_data.append(user_principal_name)
        continue
    
    # Generate the individual report for the user
    user_excel_filename = f'calls_report_{user_principal_name}_{start_date_str}_{end_date_str}.xlsx'
    generate_excel_report(user_df, user_excel_filename)
    
    # Send email to the user with their individual report
    subject = f"Your Call Data Report ({start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')})"
    body = f"Dear {user_df['userDisplayName'].iloc[0]},<br><br>Please find attached your call data report for the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}."
    send_email(subject, body, user_principal_name, [user_excel_filename])
    print(f"Individual report sent to {user_principal_name}")

# Update the email body for the manager
body = f"Dear Manager,<br><br>Please find attached the consolidated call data report for the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}."

if users_without_data:
    body += "<br><br>The following users had no call data during this period:<br>"
    body += "<br>".join(users_without_data)
else:
    body += "<br><br>All users had call data during this period."

# Send the consolidated report to the manager
subject = f"Consolidated Call Data Report ({start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')})"
send_email(subject, body, manager_email, [consolidated_excel_filename])
print(f"Consolidated report sent to the manager: {manager_email}")
